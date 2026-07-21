"""FSA つみたて投資枠対象商品リスト取込 → market.nisa_tsumitate（非破壊 upsert）。

公開データのみ・inert・規制非該当。openpyxl で 3 シートを正規化し ON CONFLICT(fund_name)。
使い方: DATABASE_URL=... FSA_XLSX=/path/nisa_shisan.xlsx [IMAJ_XLSX=...] \
        python scripts/refresh_nisa_tsumitate.py
"""
from __future__ import annotations
import datetime
import os
import sys

# シート名 → (category, fund_col, mgmt_col, index_col, df_col)。index/df 無しは None。
# col は 0 始まり列インデックス。r0-r3 メタ・r4 ヘッダ・r5 以降データ。
_SHEETS = {
    "指定インデックス投資信託": ("index", 3, 4, 2, 1),
    "指定インデックス投資信託以外の投資信託（アクティブ運用投信等）": ("active", 2, 3, None, 0),
    "上場株式投資信託（ETF）": ("etf", 1, 2, 0, None),
}
# loud-fail レンジ（実測 index286/active65/etf9）。
_RANGES = {"index": (260, 320), "active": (50, 90), "etf": (5, 15)}
_DATA_START = 5   # r0-3 メタ・r4 ヘッダ・r5 からデータ


def _fsa_list_date(ws) -> datetime.date:
    """r0 の Excel 日付シリアル（最初の数値セル）を date に変換。"""
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        if isinstance(cell, (int, float)) and cell > 40000:
            return (datetime.datetime(1899, 12, 30)
                    + datetime.timedelta(days=int(cell))).date()
    raise RuntimeError("[refresh_nisa_tsumitate] r0 date serial not found — FSA layout changed?")


def parse_fsa_tsumitate(path_or_wb) -> list[dict]:
    import openpyxl
    if isinstance(path_or_wb, str):
        wb = openpyxl.load_workbook(path_or_wb, read_only=True, data_only=True)
    else:
        wb = path_or_wb
    out: list[dict] = []
    for title, (category, fcol, mcol, icol, dcol) in _SHEETS.items():
        if title not in wb.sheetnames:
            raise RuntimeError(f"[refresh_nisa_tsumitate] missing sheet {title!r} — FSA layout changed?")
        ws = wb[title]
        list_date = _fsa_list_date(ws)
        ff = {"index_name": None, "domestic_foreign": None}   # 縦結合セルの forward-fill 状態
        n = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < _DATA_START:
                continue
            fund = row[fcol] if len(row) > fcol else None
            if not fund or not str(fund).strip():
                continue
            if icol is not None and len(row) > icol and row[icol] and str(row[icol]).strip():
                ff["index_name"] = str(row[icol]).strip()
            if dcol is not None and len(row) > dcol and row[dcol] and str(row[dcol]).strip():
                ff["domestic_foreign"] = str(row[dcol]).strip()
            mgmt = row[mcol] if len(row) > mcol and row[mcol] else None
            out.append({
                "fund_name": str(fund).strip(),
                "mgmt_company": str(mgmt).strip() if mgmt else None,
                "category": category,
                "index_name": ff["index_name"] if icol is not None else None,
                "domestic_foreign": ff["domestic_foreign"] if dcol is not None else None,
                "list_updated_at": list_date,
            })
            n += 1
        lo, hi = _RANGES[category]
        if not (lo <= n <= hi):
            raise RuntimeError(
                f"[refresh_nisa_tsumitate] {category} count {n} out of range "
                f"({lo}-{hi}) — FSA sheet {title!r} layout changed?")
    return out


import unicodedata

_TS_COLS = ("fund_name", "mgmt_company", "category", "index_name",
            "domestic_foreign", "fund_code", "etf_ticker", "list_updated_at")


def _norm_name(s) -> str:
    """全角/半角・空白差を吸収した名寄せキー。"""
    if not s:
        return ""
    return unicodedata.normalize("NFKC", str(s)).replace(" ", "").replace("　", "").lower()


def imaj_code_by_name(path_or_wb) -> dict:
    """IMAJ ファンド名称(正規化) → 5桁銘柄コード の map。"""
    import openpyxl
    if isinstance(path_or_wb, str):
        wb = openpyxl.load_workbook(path_or_wb, read_only=True, data_only=True)
    else:
        wb = path_or_wb
    ws = wb["対象商品一覧"]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        code = row[3] if len(row) > 3 else None
        name = row[4] if len(row) > 4 else None
        if code and name:
            out[_norm_name(name)] = str(code).strip()
    return out


def enrich_with_imaj(rows, path_or_wb) -> None:
    """fund_code / etf_ticker を IMAJ 名寄せで best-effort 補完（null 許容）。"""
    code_map = imaj_code_by_name(path_or_wb)
    for r in rows:
        code = code_map.get(_norm_name(r["fund_name"]))
        if code:
            r["fund_code"] = code
            if r.get("category") == "etf":
                r["etf_ticker"] = code[:4]


def upsert_tsumitate(conn, rows) -> int:
    if not rows:
        return 0
    for r in rows:                       # 欠けキーを null 埋め（IMAJ 未補完でも INSERT 可）
        r.setdefault("fund_code", None)
        r.setdefault("etf_ticker", None)
    cols = ",".join(_TS_COLS) + ",nisa_source"
    ph = ",".join([f"%({c})s" for c in _TS_COLS]) + ",'fsa-tsumitate-xlsx'"
    setexpr = ", ".join(f"{c}=EXCLUDED.{c}" for c in _TS_COLS if c != "fund_name")
    sql = (f"INSERT INTO market.nisa_tsumitate ({cols}) VALUES ({ph}) "
           f"ON CONFLICT (fund_name) DO UPDATE SET {setexpr}")
    with conn.cursor() as cur:
        cur.executemany(sql, rows)
    return len(rows)


def _connect():
    url = os.environ.get("DATABASE_URL") or os.environ.get("POSTGRES_URL")
    if not url:
        raise SystemExit("DATABASE_URL not set")
    import psycopg
    return psycopg.connect(url)


def _fetch_rows():
    path = os.environ.get("FSA_XLSX")
    if not path or not os.path.exists(path):
        raise SystemExit("FSA_XLSX not set or file missing (つみたて投資枠対象商品 xlsx)")
    rows = parse_fsa_tsumitate(path)
    imaj = os.environ.get("IMAJ_XLSX")
    if imaj and os.path.exists(imaj):
        enrich_with_imaj(rows, imaj)     # IMAJ は fund_code/etf_ticker 補完のみ（任意）
    return rows


def main(argv=None):
    argv = argv if argv is not None else sys.argv[1:]
    rows = _fetch_rows()
    with _connect() as conn:
        n = upsert_tsumitate(conn, rows)
    print(f"[refresh_nisa_tsumitate] upserted={n}", file=sys.stderr)
    return 0 if n > 0 else 1   # loud-fail: 0 件は非ゼロ


if __name__ == "__main__":
    raise SystemExit(main())
